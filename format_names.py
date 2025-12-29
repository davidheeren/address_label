#!/usr/bin/env -S uv run --script

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import sys


def get_worksheet(path: str) -> Worksheet:
    return load_workbook(path).active


def format_names(ws: Worksheet):
    for row in ws.iter_rows(min_row=2, max_row=120):
        name = str(row[0].value)
        print(name)


def main():
    ws = get_worksheet(sys.argv[1])
    format_names(ws)


if __name__ == "__main__":
    main()
