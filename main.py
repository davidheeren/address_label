#!/usr/bin/env -S uv run --script

import argparse
import webbrowser
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from pylabels import Sheet, Specification
from collections import namedtuple
from reportlab.graphics import shapes


def get_args(defaults: bool = False) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="address_label",
        description="Creates a pdf for printing address labels from an Excel file.",
    )
    parser.add_argument("-i", "--input", default="addresses.xlsx")
    parser.add_argument("-o", "--output", default="labels.pdf")
    parser.add_argument("-f", "--filter", default="*", help="Ex: 'mary joe, 4-9, !5'")
    parser.add_argument("-b", "--bias", type=int, default=0, help="Count of labels to offset")
    parser.add_argument("-n", "--name", default="", help="Your name to find return addresses row")
    parser.add_argument("-r", "--ret", action="store_true", help="Include the same number of return address labels")
    parser.add_argument("-t", "--test", action="store_true", help="Put box lines around each lablel")
    parser.add_argument("-l", "--launch", action="store_true", help="Launch the pdf in the browser")
    if defaults:
        return parser.parse_args([])
    return parser.parse_args()


def get_max_row(ws: Worksheet):
    count = 0
    for row in ws:
        if any(cell.value is not None for cell in row):
            count += 1
    return count


def match_names(ws: Worksheet, max_row: int, filter: str) -> set[int]:
    parts_a = [x.strip().lower() for x in filter.split()]
    nums = set()
    # Enumerate over data rows (skipping header if any)
    for i in range(2, max_row + 1):
        address = get_address(ws, i)
        parts_b = {
            x.strip().lower()
            for x in (
                address.last_name1,
                address.first_name1,
                address.last_name2,
                address.first_name2,
            )
            if x}
        if all(a in parts_b for a in parts_a):
            nums.add(i)
    return nums


def get_indices(ws: Worksheet, max_row: int, filter_str: str, name: str) -> (set[int], int):
    # example filter "1, 5-9, !6, John Doe"
    # note that this is dependent on each filter's order. you want to first include then exclude
    indices = set()
    filters = filter_str.split(",")

    for f in filters:
        invert = False
        nums = set()
        f = f.strip()
        if not f:
            continue
        if f.startswith("!"):
            invert = True
            f = f[1:]
        if not f:
            raise Exception("Invalid filter: dangling '!'")

        if f == "*":
            nums.update(range(2, max_row + 1))
        # filter is a name
        elif all(c.isalpha() or c.isspace() for c in f):
            match_nums = match_names(ws, max_row, f)
            nums.update(match_nums)
            print(f"Matched name: '{f}', {len(match_nums)} times")

        # filter is number or number range
        elif all(c.isdigit() or c == "-" for c in f):
            # filter is range
            if "-" in f:
                parts = f.split("-")
                if len(parts) != 2 or not parts[0].strip().isdigit() or not parts[1].strip().isdigit():
                    raise Exception(f"Invalid index range: {f}")
                start, end = map(int, (parts[0], parts[1]))
                if start < 2 or start > max_row or end < 2 or end > max_row:
                    raise Exception(f"Invalid index: {f}, out of bounds {2}-{max_row}")
                if start > end:
                    raise Exception(f"Invalid range: start > end in {f}")
                # Convert to 0-based index relative to data start
                nums.update(range(start, end + 1))
            # filter is single number
            else:
                num = int(f)
                if num < 2 or num > max_row:
                    raise Exception(f"Invalid index: {f}, out of bounds {2}-{max_row}")
                # Convert to 0-based index
                nums.add(num)

        # not a valid filter
        else:
            raise Exception(f"Not a valid filter: {f}")

        # update the indices
        if invert:
            indices.difference_update(nums)
        else:
            indices.update(nums)

    # remove name
    name_idx = -1
    if name:
        match_nums = match_names(ws, max_row, name)
        if len(match_nums) == 0:
            raise Exception(f"Name: '{name}' not found. This is needed for the return address")
        if len(match_nums) > 1:
            raise Exception(f"Name: '{name}' found multiple times. There can only be one for the return address")
        indices.difference_update(match_nums)
        name_idx = match_nums.pop()

    return indices, name_idx


def get_worksheet(input_path: str) -> Worksheet:
    path = Path(input_path)
    if not path.is_file():
        raise FileNotFoundError(f"Input file not found at: {input_path}")

    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Unsupported file type: {path.suffix}. Only .xls or .xlsx files are supported.")

    wb = load_workbook(input_path)
    return wb.active


def get_sheet(test: bool) -> Sheet:
    PADDING = 1
    specs = Specification(
        215.9,
        279.4,
        3,
        10,
        66.7,
        25.4,
        corner_radius=2,
        left_margin=5,
        right_margin=5,
        top_margin=13,
        left_padding=PADDING,
        right_padding=PADDING,
        top_padding=PADDING,
        bottom_padding=PADDING,
        row_gap=0,
    )

    if test:
        return Sheet(specs, draw_address, border=True)
    return Sheet(specs, draw_address)


Address = namedtuple(
    "Address", ["last_name1", "first_name1", "last_name2", "first_name2", "address1", "address2", "city", "state", "zip", "country"]
)


def get_address(ws: Worksheet, row: int) -> Address:
    # this functions assumes that the row is a valid index
    # row is 1 based
    values = [
        ws.cell(row=row, column=col).value
        for col in range(1, 11)
    ]
    return Address(*values)


def draw_address(label, width, height, address: Address | None):
    # this function excpects a valid address or None
    if address is None:
        return

    name = ""
    if address.last_name1 and address.first_name1 and address.last_name2 and address.first_name2:
        name = f"{address.first_name1} {address.last_name1} & {address.first_name2} {address.last_name2}"
    elif address.last_name1 and address.first_name1 and address.first_name2:
        name = f"{address.first_name1} & {address.first_name2} {address.last_name1}"
    elif address.last_name1 and address.first_name1:
        name = f"{address.first_name1} {address.last_name1}"
    elif address.last_name1:
        name = address.last_name1
    else:
        print(f"Warning: Skipping line with invalid name '{address.last_name1} {address.first_name1} {address.last_name2} {address.first_name2}'")
        return

    lines = [
        address.country,
        f"{address.city} {address.state}  {address.zip}".upper(),
        address.address1.upper(),
        address.address2.upper() if address.address2 else None,
        name,
    ]

    group = shapes.Group()
    x, y = 0, 0
    for line in lines:
        if not line:
            continue
        shape = shapes.String(x, y, line, textAnchor="start", fontSize=10)
        _, _, _, y = shape.getBounds()
        y += 3
        group.add(shape)
    _, _, lx, ly = label.getBounds()
    _, _, gx, gy = group.getBounds()

    if gx > lx:
        # raise Exception(f"Address too long, name: {address.name}")
        print(f"Warning: Address too long, name: {name}")
    if gy > ly:
        # raise Exception(f"Address too tall, name: {address.name}")
        print(f"Warning: Address too tall, name: {name}")

    dx = (lx - gx) / 2
    dy = (ly - gy) / 2
    group.translate(dx, dy)

    label.add(group)


def save_pdf(args: argparse.Namespace, ws: Worksheet, indices: set[int], sheet: Sheet, name_idx: int):
    if args.bias > 0:
        sheet.add_label(None, count=args.bias)

    for i in sorted(list(indices)):
        address = get_address(ws, i)

        if not any(address):
            print(f"Warning: Skipping blank row index '{i}'")
            continue

        if not address.last_name1 or not address.address1 or not address.city or not address.state or not address.zip:
            print(f"Warning: Skipping row with name: '{address.first_name1}', index: '{i}' due to one or more missing address fields.")
            continue

        sheet.add_label(address)

    if args.ret:
        if not args.name:
            raise Exception("Name must be set to use the ret option")
        sheet.add_label(get_address(ws, name_idx), count=len(indices))

    sheet.save(args.output)
    print(f"{sheet.label_count} label(s) output on {sheet.page_count} page(s).")


def run(args: argparse.Namespace):
    ws = get_worksheet(args.input)
    max_row = get_max_row(ws)
    indices, name_idx = get_indices(ws, max_row, args.filter, args.name)
    sheet = get_sheet(args.test)
    save_pdf(args, ws, indices, sheet, name_idx)
    if args.launch:
        webbrowser.open(args.output)


def main():
    run(get_args())
    # args = get_args()
    # ws = get_worksheet(args.input)
    # print(get_address(ws, 3))
    # try:
    #     run(get_args())
    # except Exception as e:
    #     print(f"Error: {e}")


if __name__ == "__main__":
    main()
