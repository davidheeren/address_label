
from argparse import Namespace
import webbrowser
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from pylabels import Sheet, Specification
from collections import namedtuple
from reportlab.graphics import shapes

# Pylabels2 docs example: https://github.com/erikvw/pylabels2/blob/main/pylabels/demos/addresses.py

ADDRESS_COLUMN_COUNT = 10


# Record data format
Address = namedtuple(
    "Address", ["last_name1", "first_name1", "last_name2", "first_name2", "address1", "address2", "city", "state", "zip", "country"]
)


class LabelGenerator:
    def __init__(self, args: Namespace):
        """Setup and load data"""
        self.args = args
        self.ws = self._load_worksheet()
        self.max_row = self._find_max_row()

    def _load_worksheet(self) -> Worksheet:
        """Loads an Excel file and returns the first Worksheet"""
        path = Path(self.args.input)
        if not path.is_file():
            raise FileNotFoundError(f"Input file not found at: {path}")

        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Unsupported file type: {path.suffix}. Only .xls or .xlsx files are supported.")

        wb = load_workbook(path)
        return wb.active

    def _find_max_row(self) -> int:
        """Finds the max row that is not empty, since ws.max_row counts empty rows"""
        count = 0
        for row in self.ws:
            if any(cell.value is not None for cell in row):
                count += 1
        return count

    def _get_address(self, row: int) -> Address:
        """Returns a address record from the data at the 1 based row index"""
        if row < 2 or row > self.max_row:
            raise ValueError(f"Row index: {row} out of bounds: 2-{self.max_row}")
        values = [
            self.ws.cell(row=row, column=col).value
            for col in range(1, ADDRESS_COLUMN_COUNT + 1)
        ]
        return Address(*values)

    def _split_and_format_filters(self) -> list[tuple[str, bool]]:
        """Returns a list of filters: (string, invert). Split on ','"""
        f_strs = self.args.filter.split(",")
        filters = []
        for f in f_strs:
            invert = False
            f = f.strip()
            if not f:
                continue
            if f.startswith("!"):
                invert = True
                f = f[1:]
            if not f:
                raise ValueError("Invalid filter: dangling '!'")
            filters.append((f, invert))
        return filters

    def _match_name(self, filter: str) -> set[int]:
        """
        Returns a set of all indices matched in the name fields
        All parts in the split input filter, must match an address name field
        """
        filter_parts = [x.strip().lower() for x in filter.split()]
        indices = set()
        for i in range(2, self.max_row + 1):
            address = self._get_address(i)
            # Address name fields if they exist
            address_parts = {
                x.strip().lower()
                for x in (
                    address.last_name1,
                    address.first_name1,
                    address.last_name2,
                    address.first_name2,
                )
                if x}
            if all(a in address_parts for a in filter_parts):
                indices.add(i)
        return indices

    def _match_index_or_range(self, filter: str) -> set[int]:
        """Determines if filter is index or range and returns set of the matched indices"""
        # Filter is a range
        if "-" in filter:
            parts = filter.split("-")
            if len(parts) != 2 or not parts[0].strip().isdigit() or not parts[1].strip().isdigit():
                raise ValueError(f"Invalid index range: {filter}")
            start, end = map(int, (parts[0], parts[1]))
            if start < 2 or start > self.max_row or end < 2 or end > self.max_row:
                raise ValueError(f"Invalid index: {filter}, out of bounds {2}-{self.max_row}")
            if start > end:
                raise ValueError(f"Invalid range: start > end in {filter}")
            return set(range(start, end + 1))

        # Filter is single number
        num = int(filter)
        if num < 2 or num > self.max_row:
            raise ValueError(f"Invalid index: {filter}, out of bounds {2}-{self.max_row}")
        return {num}

    def _filter_indices(self) -> tuple[set[int], int]:
        """
        Finds all matched indices from the filter argument
        Also removes the name input arg
        Returns the set of matched indices and the name index
        """
        filters = self._split_and_format_filters()
        # Result indices to return
        indices = set()
        for filter, invert in filters:
            # Set of nums to either add to or remove from indices
            nums = set()

            # Filter is wildcard
            if filter == "*":
                nums.update(range(2, self.max_row + 1))

            # Filter is a name
            elif all(c.isalpha() or c.isspace() for c in filter):
                match_nums = self._match_name(filter)
                nums.update(match_nums)
                print(f"Matched name: '{filter}', {len(match_nums)} times")

            # Filter is number or number range
            elif all(c.isdigit() or c == "-" for c in filter):
                match_nums = self._match_index_or_range(filter)
                nums.update(match_nums)

            # Not a valid filter
            else:
                raise ValueError(f"Not a valid filter: {filter}")

            # Update the indices based on the invert
            if invert:
                indices.difference_update(nums)
            else:
                indices.update(nums)

        # Remove name
        name_idx = -1
        if self.args.name:
            match_nums = self._match_name(self.args.name)
            if len(match_nums) == 0:
                raise ValueError(f"Name: '{self.args.name}' not found. This is needed for the return address")
            if len(match_nums) > 1:
                raise ValueError(f"Name: '{self.args.name}' found multiple times. There can only be one for the return address")
            indices.difference_update(match_nums)
            name_idx = match_nums.pop()

        return indices, name_idx

    # Makes the method not take self as a parameter
    @staticmethod
    def _draw_address(label, width, height, address: Address | None):
        """Draws an address to a label"""
        if address is None:
            return

        # Formats the name based on which entries are empty or not
        # John Miller, John & Mary Miller, John Miller & Mary Sue
        # If a single name, like a company, should only have last_name1
        name = ""
        if address.last_name1 and address.first_name1 and address.last_name2 and address.first_name2:
            name = f"{address.first_name1} {address.last_name1} & {address.first_name2} {address.last_name2}"
        elif address.last_name1 and address.first_name1 and address.first_name2:
            name = f"{address.first_name1} & {address.first_name2} {address.last_name1}"
        elif address.last_name1 and address.first_name1:
            name = f"{address.first_name1} {address.last_name1}"
        elif address.last_name1:
            name = address.last_name1
        else:
            print(f"Warning: Skipping line with invalid name '{address.last_name1} {address.first_name1} {address.last_name2} {address.first_name2}'")
            return

        # Only include the PO box, if not empty
        street = address.address1
        if address.address2:
            street = address.address2

        # Lines of the label
        # Uppercase for maximal readability
        lines = [
            address.country.upper() if address.country else None,
            f"{address.city} {address.state}  {address.zip}".upper(),
            street.upper(),
            name,
        ]

        # From the pylables2 library docs
        group = shapes.Group()
        x, y = 0, 0
        for line in lines:
            if not line:
                continue
            shape = shapes.String(x, y, line, textAnchor="start", fontSize=10)
            _, _, _, y = shape.getBounds()
            y += 3
            group.add(shape)
        _, _, lx, ly = label.getBounds()
        _, _, gx, gy = group.getBounds()

        if gx > lx:
            print(f"Warning: Address too long, name: {name}")
        if gy > ly:
            print(f"Warning: Address too tall, name: {name}")

        dx = (lx - gx) / 2
        dy = (ly - gy) / 2
        group.translate(dx, dy)

        label.add(group)

    def _create_sheet(self) -> Sheet:
        """Creates a sheet that can be saved as a PDF"""
        # Follows the Avery 8160 specs
        # From the pylabels2 docs
        padding = 1
        specs = Specification(
            215.9,
            279.4,
            3,
            10,
            66.7,
            25.4,
            corner_radius=2,
            left_margin=5,
            right_margin=5,
            top_margin=13,
            left_padding=padding,
            right_padding=padding,
            top_padding=padding,
            bottom_padding=padding,
            row_gap=0,
        )
        if self.args.test:
            return Sheet(specs, self._draw_address, border=True)
        return Sheet(specs, self._draw_address)

    def _save_pdf(self, sheet: Sheet, indices: set[int], name_idx: int):
        """Saves the sheet as a PDF"""
        # Add blank labels
        if self.args.bias > 0:
            sheet.add_label(None, count=self.args.bias)

        # Add labels for the indices
        for i in sorted(list(indices)):
            address = self._get_address(i)

            if not any(address):
                print(f"Warning: Skipping blank row index '{i}'")
                continue

            if not address.last_name1 or not address.address1 or not address.city or not address.state or not address.zip:
                print(f"Warning: Skipping row with name: '{address.first_name1}', index: '{i}' due to one or more missing address fields.")
                continue

            sheet.add_label(address)

        # Add return address labels
        if self.args.ret:
            if not self.args.name:
                raise ValueError("Name must be set to use the ret option")
            sheet.add_label(self._get_address(name_idx), count=len(indices))

        sheet.save(self.args.output)
        print(f"{sheet.label_count} label(s) output on {sheet.page_count} page(s).")

    def generate_pdf(self):
        indices, name_idx = self._filter_indices()
        sheet = self._create_sheet()
        self._save_pdf(sheet, indices, name_idx)
        if self.args.launch:
            webbrowser.open(self.args.output)
